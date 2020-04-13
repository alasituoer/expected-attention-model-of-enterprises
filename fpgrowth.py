import pandas as pd
import numpy as np

class treeNode:

    def __init__(self, nameValue, numCount, parentNode):
        """
            初始化结点时传入名称, 统计值, 父结点
            name: 结点名(元素名)
            count: 统计值
            nodeLink: 链接相似的元素项
            parent: 父结点
            children: 孩子结点
        """

        self.name = nameValue
        self.count = numCount
        self.nodeLink = None
        self.parent = parentNode
        self.children = {}

    def inc(self, numOccur):
        """给统计值加上一个指定值"""

        self.count += numOccur

    def disp(self, ind=0):
        """从根结点开始打印数结构"""

        print('    '*ind, '%s:%s'% (self.name, self.count))
        for child in self.children.values():
            child.disp(ind+1)

class fpGrowth:

    def __init__(self, df_train, minSup, minConf, init_weight=False):
    
        self.df_train = df_train
        self.numItems = df_train.shape[0]
        self.minSup = minSup
        self.minConf = minConf
        self.init_weight = init_weight

        # 设置最小支持度阈值, 如果传入的是小数, 则认为是占比, 将其转换为数量
        if self.minSup < 1:
            self.minSup = self.minSup * self.numItems
        #print('minSup', self.minSup)

    def loadDatasetsSample1(self):
        """生成样本数据1"""

        self.datasets_init = [
                ['r', 'z', 'h', 'j', 'p',],
                ['z', 'y', 'x', 'w', 'v', 'u', 't', 's',],
                ['z',],
                ['r', 'x', 'n', 'o', 's',],
                ['y', 'r', 'x', 'z', 'q', 't', 'p',],
                ['y', 'z', 'x', 'e', 'q', 's', 't', 'm',],]

    def loadDatasetsSample2(self):
        """生成样本数据2"""

        chaosList = [chr(x) for x in range(ord('A'), ord('F')+1)]
            #+\[chr(x) for x in range(ord('a'), ord('z')+1)] + [*range(10)] 
        np.random.seed(6143)
        self.datasets_init = np.random.choice(chaosList, (1000, 10)).tolist()

    def loadDatasetsSample3(self):
        """生成样本数据3"""

        self.datasets_init = [
                ['li', 'ming', 'zhi'],
                ['alas', 'tuoer', 'stain', 'ein'],
                ['li', 'zhi', 'ming', 'li', 'zhi'],
                ['li', 'alas', 'stain'],
                ['li', 'tuoer', 'ein'],
                ['ein', 'ming', 'alas'],]
                #['li', 'zhi', 'tuoer']]

    
        return self.datasets_init

    def loadWeightSample(self):
        """生成样本数据3的项集权重"""

        self.df_weight = pd.DataFrame(
                [['li', 5], ['ming', 2], ['zhi', 5], ['alas', 4],
                ['tuoer', 5], ['stain', 1], ['ein', 4]], 
                columns=['feature_group_name', 'weight'])

    def loadAndTransDatasetsReal(self):
        """
            读入样本数据(随机生成的模拟真实数据)
            给连续特征分组, 返回修改后的新样本数据
            同时可输出特征分组文件, 供手动修改分组权重值
        """
    
        df = self.df_train.copy()
    
        # 1 特征转换: 指标值分组和转换
        # 1.1 对类别特征, 用[特征名+'__'+特征值]的结果替换原来的特征值
        # 如将'建筑业'替换成'industry__建筑业'
        for f1 in ['industry', 'city', 'district', 'is_change_premises', 
                'cacellation_reason', 'is_real_premise',
                'is_executee', 'is_financial_black_list', 'alert_level'] :
            df[f1] = df[f1].map(lambda x: f1 + '__' + str(x))
            #print(df[f1].value_counts())
    
        # 1.2 对连续特征值, 先分组, 再用[指标名+'__'+分组]的结果替换分组后的指标值
        for f2 in ['registered_capital', 'employees', 'total_assets', 
                'total_tax', 'industry_index', 'fraud_score',
                'period_abnormal', 'rate_frozon_holdings', 'total_liabilities', 
                'total_debt', 'judical_auction_amount', 'complaints_number_monthly',
                'rate_conciliation', 'sub_enterprises_number',]:
            df[f2] = pd.cut(df[f2], 4, duplicates='drop').map(
                    lambda x: f2 + '__' + str(x))
            #print(df[f2].value_counts())
    
        # 1.3 给个别指标手动指定分组后改名
        #print(df['period_abnormal'].value_counts())
        #print(pd.cut(df['period_abnormal'], bins=5))
    
        del df['enterprise_id']
    
        # 2 输出转换后的数据集
        # 2.1 输出所有指标的所有分组值, 存储在features_group_name.xlsx文件中
        df_features_group = pd.DataFrame()
        for c in df.columns:
            #print(c)
            df1 = df[c].value_counts().reset_index()
            df1.columns = ['feature_group_name', 'counts']
            df1['interval'] = df1['feature_group_name'].map(lambda x: x.split('__')[1])
            #print(df1[['features', 'interval', 'counts']])
            df_features_group = pd.concat([df_features_group, 
                    df1[['feature_group_name', 'interval', 'counts']]])
        #print(df_features_group)
        df_features_group.to_excel('datasets/features_group_name.xlsx', index=False)
    
        # 2.2 存储分组转换后的指标值
        #print(df.shape, '\n', df.head(), '\n', df.columns)
        df.to_csv('datasets/features_grouped.csv', sep=';', index=False)
    
        # 3 将特征转换后的数据集转换成数组格式, 模型利用datsets做正式训练
        self.datasets_init = df.to_numpy().tolist()

        return self.datasets_init

    def loadWeightReal(self):
        """读入上述模拟真实数据的项集权重"""

        self.df_weight = pd.read_excel('datasets/features_group_name_adjusted.xlsx')

    def initDatasets(self, datasets):
        """将每条记录转换为集合格式, 以组合为键, 组合计数为值构造字典返回"""
    
        dict_set = {}
        for trans in datasets:
            dict_set[frozenset(trans)] = dict_set.get(frozenset(trans), 0) + 1
    
        return dict_set
    
    def createTree(self, datasets, dict_freq_pre, dict_freq_sup):
        """
            构建FP树或条件FP树, 查找频繁项集, 
			首次创建FP树时发现的是单元素频繁项集, 
			创建条件FP树时发现的是多元素频繁项集.
            输入:
                datasets: 经过预处理的训练数据集, frozenset(每条样本)的列表格式
                dict_freq_pre: 样例{}或{'', ''}, 
                    如果用于构造FP树, 传入的是空集合{}
                    如果用于构造条件FP树, 传入的是已组合好的频繁项集
                dict_freq_sup: 保存创建FP树和条件FP树时发现的频繁项集
            输出:
                fpTree: 所有记录的FP树
                dict_header: 单元素频繁项集及其总计数, 指向该元素的第一个实例的指针
                    存储格式: {name: [count, nodeLink], ...}
            其他:
                同时对dict_freq_sup也做了更新, 分别在创建FP树和条件FP树时操作
        """
    
        #1 第一次遍历数据集, 统计单元素的频数, 存放在头指针表
        dict_header = {}
        for trans in datasets:
            # 存在则将该项的计数值提取并加1(=值datesets[trans])存储
            for item in trans:
                dict_header[item] = dict_header.get(item, 0) + datasets[trans]
        #print(dict_header)
    
        #2 根据是否给项集增加权重而分路经处理
        if self.init_weight:
            # 后面对字典做了更改, 所以此处的循环先显示地复制迭代序列为列表
            # RuntimeError: dictionary changed size during iteration
            for k in [*dict_header.keys()]:
                # 获取当前项集的权重值, 这里得分情况
                # 1.创建FP树时当前项集即为单元素项集k, 此处weight即为k的权重
                # 2.创建条件FP树时当前项集为传入的dict_freq_pre加上k, 
                #   此处的weight是list(dict_freq_pre)+[k]元素中的最小权重
                weight = self.df_weight[self.df_weight['feature_group_name'
                        ].isin(list(dict_freq_pre)+[k])]['weight'].min()
                # 更新头指针表, 这里没有对FP树或条件FP树的结点计数做更改
                dict_header[k] = dict_header[k] * weight
                # 如果头指针表中的元素支持度不满足最小支持度, 
                # 则说明该项集不满足最小支持度要求, 删除该元素
                if dict_header[k] < self.minSup:
                    del dict_header[k]
                # 否则说明该项集是频繁项集, 保存到dict_freq_sup
                else:
                    #print(dict_freq_pre, k, dict_header[k])
                    d = frozenset(list(dict_freq_pre)+[k])
                    #print(d)
                    dict_freq_sup.update({d: dict_header[k]})
        # 如果不修改项集权重, 就根据头指针表的计数直接判断是否频繁
        # 同时也保存频繁项集到dict_freq_sup
        else:
            for k in [*dict_header.keys()]:
                if dict_header[k] < self.minSup:
                    del dict_header[k]
                else:
                    #print(dict_freq_pre, k, dict_header[k])
                    d = frozenset(list(dict_freq_pre)+[k])
                    #print(d)
                    #print(dict_freq_pre.update(k), dict_header[k])
                    dict_freq_sup.update({d: dict_header[k]})
        # 或者按照下述路径分开处理, 上面将对FP树和条件FP树传入的项集做了合并
        #if self.init_weight: and not list_cond[0]:
        #elif self.init_weight and list_cond[0]:
        #else:

        set_freq = set(dict_header.keys())
        #print(set_freq)
    
        #3 如果头指针表中元素为空, 则提前返回
        if len(set_freq) == 0: return None, None
        # 扩展头指针表, 用于保存计数值和指向每种类型第一个元素项的指针
        for k in dict_header:
            dict_header[k] = [dict_header[k], None]
    
        # 第二次遍历数据集, 利用频繁项构建FP树
        fpTree = treeNode('Null Set', 1, None)
        # 对每条样本记录(count多数为1, 部分值可能大于1)
        for transet, count in datasets.items():
            #print(transet, count)
            # 构建元素的总统计值字典
            localD = {}
            for item in transet:
                if item in set_freq:
                    localD[item] = dict_header[item][0]
            #print(localD)

            # 如果该字典中有频繁项集
            if len(localD) > 0:
                # 则按照 (先)计数值,(后)元素名 降序排序成列表
                list_ordered = [v[0] for v in sorted(localD.items(), 
                        key=lambda x: (x[1], x[0]), reverse=True)]
                #print(list_ordered)
                # 并用该列表构建FP树
                self.updateTree(list_ordered, fpTree, dict_header, count)
            #print('\n')
        #print(dict_header)

        return fpTree, dict_header

    def addWeightForFpTree(self, fpTree, dict_header):
        """
            已废弃.
            更新fpTree中结点的计数值, 乘以指定权重, fpTree可不传入
            只更新fpTree中的结点, dict_header中的总计数值在createTree时已修改
        """

        for k in [*dict_header.keys()]:
            weight = self.df_weight[self.df_weight[
                    'feature_group_name']==k]['weight'].iloc[0]

            # 更新fpTree, 沿着dict_header给每个元素实例乘以给定权重
            nodeStart = dict_header[k][1]
            while(nodeStart.nodeLink != None):
                nodeStart.count = nodeStart.count * weight
                nodeStart = nodeStart.nodeLink
            # 指向最后一个实例后, 再修改最后一个实例的计数值
            nodeStart.count = nodeStart.count * weight
                #print(nodeStart.name, nodeStart.count)

        return fpTree, dict_header
    
    def updateTree(self, list_items, fpTree, dict_header, count):
        """
            功能: 被createTree调用, 将排好序的某条记录元素序列更新到已有FP树
            输入: 
                list_items: 按(计数值, 名称)排好序的某条记录的元素序列
                fpTree: 根据上条记录创建的FP树(已有FP树)
                dict_header: 传入FP树中元素项的总计数
                count: 该元素项出现记录数, 总计数应增加数
            输出: 无, 实际上是对fpTree和dict_header做了更新
        """

        # 取到该序列中计数最多的元素
        item = list_items[0]
    
        # 检测列表中的第一项是否在子结点中, 如果在则计数加一
        if item in fpTree.children:
            fpTree.children[item].inc(count)
        # 否则创建新分支存储新结点, 并更新头指针指向新的结点
        else:
            # 新建结点时指定了新结点的父结点parent
            fpTree.children[item] = treeNode(item, count, fpTree)
            # 如果dict_header没有保存这种类型第一个元素的指针
            if dict_header[item][1] == None:
                # 则保存
                dict_header[item][1] = fpTree.children[item]
            # 否则, 从头指针表的nodeLink开始, 
            # 一直沿着nodeLink链接该元素的每一个实例到末尾
            else:
                self.updateHeader(dict_header[item][1], fpTree.children[item])
    
        #fpTree.disp()
        #print(dict_header)
    
        # 递归调用自身处理序列中(计数次多)的后续元素
        if len(list_items) > 1:
            self.updateTree(list_items[1:], fpTree.children[list_items[0]], 
                    dict_header, count)
    
    def updateHeader(self, nodeToTest, targetNode):
        """
			被updateTree调用, 更新头指针表链接到树中新的结点
            确保结点链接nodeLink指向树中该元素项的每一个实例
        """
    
        while(nodeToTest.nodeLink != None):
            nodeToTest = nodeToTest.nodeLink
        nodeToTest.nodeLink = targetNode
    
    
    def mineTree(self, fpTree, dict_header, preFix, dict_freq_sup):
        """
            功能: 创建条件树, 发现前缀路径和条件基, 来挖掘频繁项集
            输入:
                fpTree: 构建好的FP树
                dict_header: FP树中元素的头指针表, 
                        包括元素的总计数和相似元素的第一个实例
                self.minSup: 最小支持度
                dict_freq_the_item: 存储某个元素项与其条件基内元素组合成的频繁项集,
                        每次递归都可能会被被扩充(更新), 初次传入时为空
                dict_freq_sup: 所有频繁项集的支持度
            输出: 无, 实际上是不断递归时对list_freq_all和dict_freq_sup直接做了修改
            其他:
                后添加的参数dict_freq_sup是为了存储频繁项集的支持度
                但同时也保存有频繁项集, 与传入list_freq_all有重复, 
                现没有传入list_freq_all但是保留了写法, 注释了
        """
    
        # 对头指针表内的元素项按照统计值升序排序, 生成列表
        L = [x[0] for x in sorted(dict_header.items(), key=lambda k: k[1][0])]
        #print('\nL', L)
    
        # 对每个元素
        for item in L:
            #print('当前元素: ', item)
    
            # 更新当前元素与其条件基元素组合的频繁项集dict_freq_the_item, 首次传入时为空
            dict_freq_the_item = preFix.copy()
            # 当上一级的条件基中有满足最小支持度的元素组合时, 
            # 会递归组合历史频繁元素与条件基中的元素
            dict_freq_the_item.add(item)
            #print('当前元素的频繁项集: dict_freq_the_item = ', dict_freq_the_item)
            # 将各种元素的频繁项集追加到最终的频繁项集列表
    
            # 递归获取该元素的条件基, 传入待查询元素和该元素的第一个实例
            condPattBases = self.findPrefixPath(item, dict_header[item][1])
            #print('该元素的条件模式基: condPattBases = ', condPattBases)
            # 根据条件基创建条件树(树中的元素都大于最小支持度)
            condTree, condHead = self.createTree(condPattBases, 
                    dict_freq_the_item, dict_freq_sup)
            #print('条件头指针表: condHead = ', condHead)
    
            # 如果树中有元素项的话, 递归调用自己, 直到条件树中只有一个元素
            # 在此过程不断更新该元素的频繁项集到dict_freq_the_item
            if condHead != None:
                #print('conditional tree for: ', dict_freq_the_item)
                #condTree.disp()
    
                self.mineTree(condTree, condHead, dict_freq_the_item, dict_freq_sup)

    def ascendTree(self, leafNode, prefixPath):
        """
            被self.findPrefixPath()调用, 获取前缀路径
            递归访问给定结点的父结点, 直到根结点
            获取结点的名称存入(直接修改)prefixPath
        """
    
        if leafNode.parent != None:
            prefixPath.append(leafNode.name)
            self.ascendTree(leafNode.parent, prefixPath)
    
    def findPrefixPath(self, item, treeNode):
        """
            功能: 被self.mineTree()调用, 获取指定元素项的条件模式基
            输入:
                item: 待查询元素项
                treeNode: 该元素项在头指针表中的第一个实例
            输出:
                condPats: 条件基集合
            其他: 传入的item在当前函数没有直接用到, 
                它是用于在头指针表中做提取
        """
    
        # 初始化该元素项的条件模式基存储结构
        condPats = {}
        # 从头指针表中该元素项的第一个实例开始, 查找条件模式基(前缀路径)
        while treeNode != None:
    
            # 传入空的prefixPath用于存储前缀路径
            prefixPath = []
            # 在ascendTree中直接修改了prefixPath
            self.ascendTree(treeNode, prefixPath)
            # 如果前缀路径长度大于1, 才存储
            if len(prefixPath) > 1:
                condPats[frozenset(prefixPath[1:])] = treeNode.count
    
            # 由结点链接指向该相似元素项的下一个实例, 即移向链表的下一项
            treeNode = treeNode.nodeLink
    
        return condPats
    
    def computeSupData(self, xSet, fpTree, headerDict):
        """
            未使用, 计算指定频繁项集的支持度
            输入:
                xSet: 指定频繁项集
                fpTree: 根据原始数据集构建的FP树
                headerDict: 头指针表
            输出:
                xListSorted: 按照建树时元素顺序排好的组合元素列表
                supData: 计算好的支持度
        """
    
        #fpTree.disp()
    
        # 如果传入的只有一个元素, 则直接从头指针表中查找返回
        if len(xSet) == 1:
            #item = list(xSet)[0]
            item = xSet.pop()
            return [item], headerDict[item][0]
    
        # 不止一个元素, 则先按照构建FP树时的头指针表计数大小升序排列
        xSetTrans = {x: headerDict[x][0] for x in xSet}
        xList = sorted(xSetTrans.items(), key=lambda x: (x[1], x[0]))
        xListSorted = [x[0] for x in xList]
        #print('排好序的待查询组合:', xListSorted)
    
        # 查找计数值最小的元素的条件基
        item = xListSorted[0]
        condBases = self.findPrefixPath(item, headerDict[item][1])
        #print('%s的前缀路径:'% item, condBases)
    
        # 判断有哪些条件基包含剩余的所有其他元素
        supDataList = []
        for k, v in condBases.items():
            flagList = []
            for i in xListSorted[1:]:
                if i in k:
                    flagList.append(1)
                else:
                    flagList.append(0)
            #print(k, v, flagList)
    
            # 挑选出满足要求的条件基, 保存该路径上的值
            #    all([])竟然等于Ture, 不得已再加上条件长度大于0
            if (len(flagList) > 0) and all(flagList):
                supDataList.append(v)
        #print('路径上的支持度', supDataList)
    
        '''
        if self.init_weight:
            weight = self.df_weight[self.df_weight['feature_group_name'
                    ].isin(list(xSet))]['weight'].min()
            supData = sum(supDataList) * weight
        else:
            supData = sum(supDataList)
        '''

        # 满足要求的路径可能有多条, 取和作为该组合的支持度返回
        return xListSorted, sum(supDataList)
    
    def generateRules(self, L, fpTree, dict_header):
        """
            功能: 利用频繁项集挖掘关联规则
            输入: 
                L: 所有的频繁项集, 
                fpTree, dict_header: 计算置信度所需参数
                self.minConf: 置信度阈值, 默认为0.5
            输出:
                list_rules_all: 关联规则列表的列表
        """
        #print(L)
        #fpTree.disp()
    
        # 初始化存储关联规则的列表
        list_rules_all = []
        # 单个元素的频繁项集无法产生关联规则, 所以从两个元素的L[1]开始
        for i in range(1, len(L)):
            # 初始化存储不同元素数的规则列表
            list_rules_one_item = []
            # 对每个频繁项集
            for freqSet in L[i]:
                # H1和freqSet的关系
                # H1表示某个频繁项集中的单元素固定集合列表
                H1 = [set([item]) for item in freqSet]
                #print('freqSet:\t', freqSet)
                #print('H1:\t', H1)
    
                # 如果是由两个以上元素组成的频繁项集
                if i > 1:
                    # 则通过最初的项集构建更多的关联规则, 再判断关联规则的置信度
                    self.rulesFromConseq(freqSet, H1, fpTree, dict_header, 
                            list_rules_one_item)
                # 否则对于两个元素的项集,
                # 直接计算当前项集的关联规则的置信度, 挑选关联规则
                else:
                    self.calcConf(freqSet, H1, fpTree, dict_header, 
                            list_rules_one_item)
                #print('\n')
            list_rules_all.append(list_rules_one_item)
    
        # 上述for循环中传入的list_rules_one_item, 
        # 会在rulesFromConseq()和calcConf()中更新
        return list_rules_all
    
    def calcConf(self, freqSet, H, fpTree, dict_header, brl):
        """
            功能: 计算指定前后件规则的置信度
            输入:
                freqSet: 频繁项集
                H: 频繁项集中的单元素固定集合列表
                fpTree, dict_header: 计算置信度所需参数
                brl: 规则列表
                self.minConf: 允许的最小置信度
            输出: 
                prunedH: 满足最小置信度的规则列表
            其他: 实际上还使用了频繁项集的支持度数据self.dict_freq_sup
        """
    
        # 初始化满足最小可信度minConf的后件规则列表
        prunedH = []
    
        # 依次以H中的每项作为后件规则, 剩余项作为前件规则, 构造关联规则
        # 只有两个元素项, 即使其中一个指向另一个是关联规则, 但是反过来不一定是
        for conseq in H:
            #xList1, supData1 = self.computeSupData(freqSet, fpTree, dict_header)
            #xList2, supData2 = self.computeSupData(freqSet-conseq, fpTree, dict_header)
            supData1 = self.dict_freq_sup[frozenset(freqSet)]
            supData2 = self.dict_freq_sup[frozenset(freqSet-conseq)]
            conf = supData1/supData2
            #print('conseq:\t', conseq)
            #print('freqSet-conseq:\t', freqSet-conseq)
            #print('conf:\t', conf)
    
            # 筛选出满足最小置信度的规则
            if conf >= self.minConf:
                #print(freqSet-conseq, '-->', conseq, 'conf:', conf)
                brl.append((freqSet-conseq, conseq, round(conf, 4)))
                # 后件规则存入prunedH中
                prunedH.append(conseq)
        #print(prunedH)
    
        return prunedH
    
    def rulesFromConseq(self, freqSet, H, fpTree, dict_header, brl):
        """
            功能: 从最初的项集生成更多的关联规则, 判断是否满足最小置信度
            输入:
                freqSet: 频繁项集
                H: 频繁项集中的单元素固定集合列表
                dict_headr: 头指针表
                brl: 规则列表
                self.minConf: 允许的最小置信度
            输出: .
        """
    
        m = len(H[0])
        if len(freqSet) > (m+1):
            # 生成H中元素的无重复组合, Hmp1会替代H作为后续的后件规则接受检测
            Hmp1 = self.aprioriGen(H, m+1)
            #print('H:', H)
            #print('Hmp1:', Hmp1)
            # 之后Hmp1中的后件规则的元素数会大于等于2
            # 测试Hmp1中元素作为后件规则构成的关联规则是否满足最小置信度
            Hmp1 = self.calcConf(freqSet, Hmp1, fpTree, dict_header, brl)
            # 如果不止一条规则满足要求, 则继续迭代判断是否进一步组合规则
            if len(Hmp1) > 1:
                self.rulesFromConseq(freqSet, Hmp1, fpTree, dict_header, brl)
    
    def aprioriGen(self, Lk, k):
        """根据传入的k元频繁项集, 生成k+1元候选项集"""
    
        retList = []
        lenLk = len(Lk)
        for i in range(lenLk):
            for j in range(i+1, lenLk):
                # 注: 原书中对Lk[i]取值后先进行排序再截取的方法不对, 
                # list(frozenset)得到的列表中元素顺序会发生变化
                L1 = list(Lk[i])
                L2 = list(Lk[j])
                L1.sort()
                L2.sort()
                # 仅需要比较k-2个(k从2开始), 即不用比较排序后的最后一位
                # 两组候选项集的前k-1个元素相同, 之后最后一位不同
                # 求集合并集后有k+1个元素, 达成目的(不用循环列表来寻找非重复值)
                if L1[:k-2] == L2[:k-2]:
                    retList.append(Lk[i]|Lk[j])
    
        return retList
    
    def saveModel(self, filename, list_freq, fpTree, dict_header, 
            list_rules_all, list_rules_sel):
        """
            功能: 存储输入数据到指定文件
            输入:
                filename: 指定的存储文件名, excel文件
                list_freq: 频繁项集
                fpTree: FP树
                dict_header: 头指针表
                list_rules_all: 多元项集关联规则列表的列表
                list_rules_sel: 筛选后的关联规则列表
        """
    
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '频繁项集'
        for l in list_freq:
            for i in [list(x) for x in l]:
                #print(i)
                ws.append([';'.join(i)])
    
        ws1 = wb.create_sheet('头指针表')
        ws1.append(['集合项', '支持度'])
        #print([*dict_header.items()][0])
        for i in dict_header.items():
            ws1.append([i[0],i[1][0]])
    
        ws2 = wb.create_sheet('关联规则')
        #print(rules[-1])
        ws2.append(['前件规则', '后件规则', '可信度'])
        for l in list_rules_all:
            for r in l:
                #print([';'.join(list(r[0])), ';'.join(list(r[1])), r[2]])
                ws2.append([';'.join(list(r[0])), ';'.join(list(r[1])), r[2]])
    
        ws3 = wb.create_sheet('需关注的关联规则')
        ws3.append(['前件规则', '后件规则', '可信度'])
        for r in list_rules_sel:
            ws3.append([';'.join(list(r[0])), ';'.join(list(r[1])), r[2]])
    
        wb.save(filename)
    
def trainModel(df_train, minSup, minConf, init_weight, sampleOrReal):
    """模型训练主函数"""

    #1 初始化fpGrowth处理类
    fp = fpGrowth(df_train, minSup, minConf, init_weight)

    #2 读入训练数据
    if sampleOrReal:
        #datasets_init = fp.loadDatasetsSample1()
        #fp.loadWeightSample1()
        #datasets_init = fp.loadDatasetsSample2()
        #fp.loadWeightSample2()
        datasets_init = fp.loadDatasetsSample3()
        fp.loadWeightSample()
    else:
        datasets_init = fp.loadAndTransDatasetsReal()
        fp.loadWeightReal()
    #print(len(datasets_init), '\n', datasets_init[:1])
    #print('df_weight', df_weight)

    #3 转换原始数据集为字典格式(包含初步统计值)
    datasets = fp.initDatasets(datasets_init)

    #4 创建FP树, 并返回FP树和头指针表, 同时存储单元素频繁项集及其支持度
    dict_freq_sup = {}
    fpTree, dict_header = fp.createTree(datasets, {}, dict_freq_sup)
    print('FP树:')
    #fpTree.disp()
    print('\n头指针表:')
    if len(dict_header) > 0:
        for i,j in dict_header.items():
            print(i, j[0])

    #5 获取频繁项集, 同时保存多元素频繁项集的支持度数据(本段代码已废弃已注释)
    #list_freq_all = []
    #fp.mineTree(fpTree, dict_header, set([]), list_freq_all, dict_freq_sup)
    #5 按元素数长度将频繁项集分组
    #list_freq = []
    #for i in range(max([len(x) for x in list_freq_all])):
    #    list_freq.append([x for x in list_freq_all if len(x)==i+1])

    #5 获取多元素频繁项集及其支持度数据, 存储在dict_freq_sup中
    fp.mineTree(fpTree, dict_header, set([]), dict_freq_sup)
    list_freq_all = dict_freq_sup.keys()
    list_freq = []
    for i in range(max([len(x) for x in list_freq_all])):
        list_freq.append([x for x in list_freq_all if len(x)==i+1])
    #print('有支持度字典计算的频繁项集列表:', [len(x) for x in list_freq])
    print('\n频繁项集数:', len(list_freq_all))
    print('按长度分类的频繁项集数:', [len(x) for x in list_freq])
    #print('包含所有项集的列表:')
    #for l in list_freq:
    #    for it in l:
    #        print(it, fp.computeSupData(it.copy(), fpTree, dict_header)[1])


    #6 将得到的所有频繁项集的支持度保存到fp对象中, 供挖掘关联规则时计算置信度
    fp.dict_freq_sup = dict_freq_sup

    #7 挖掘关联规则
    list_rules_all = fp.generateRules(list_freq, fpTree, dict_header)
    rules_count = [len(l) for l in list_rules_all]
    print('\n关联规则数:', sum(rules_count), rules_count)
    #print(list_rules_all)
    #for l in list_rules_all:
    #    for k in l:
    #        print(k[0], k[1], k[2])

    #8 筛选关注规则
    list_rules_sel = [x for y in list_rules_all for x in y 
            if 'alert_level__5' in x[1] or 'alert_level__4' in x[1]]
    print('\n关注规则数:', len(list_rules_sel))

    #9 保存模型结果
    #'''
    filename = 'datasets/rules.xlsx'
    fp.saveModel(filename, list_freq, fpTree, dict_header, 
            list_rules_all, list_rules_sel)
    print('\n模型结果已保存')
    #'''


def testComputingSupData(df_train, minSup, minConf, init_weight):

    fp = fpGrowth(df_train, minSup, minConf, init_weight)

    # 读入训练数据
    datasets_init = fp.loadDatasetsSample()
    #datasets_init = fp.loadAndTransDatasetsReal()
    #print(len(datasets_init), '\n', datasets_init[:1])

    # 转换原始数据集为字典格式(包含初步统计值)
    datasets = fp.initDatasets(datasets_init)

    #1 创建FP树, 并返回FP树和头指针表
    fpTree, dict_header = fp.createTree(datasets, dict_freq_pre={})
    #print('#1 FP树:')
    #fpTree.disp()
    print('\n#1 头指针表:')
    for i,j in dict_header.items():
        print(i, j[0])

    #2 获取频繁项集
    list_freq_all = []
    fp.mineTree(fpTree, dict_header, set([]), list_freq_all)
    list_freq = []
    for i in range(max([len(x) for x in list_freq_all])):
        list_freq.append([x for x in list_freq_all if len(x)==i+1])
    print('\n#2 频繁项集数:', len(list_freq_all))
    #print('包含所有项集的列表:', list_freq_all)
    print('按长度分类的频繁项集数:', [len(x) for x in list_freq])

    # 比较两种方法计算得到的支持度
    print('#3 支持度计算')
    ss = pd.Series(datasets_init).map(set)
    #sup_item = ss.map(lambda x: 1 if {'E'}.issubset(x) else 0).sum()
    #print(ss, sup_item)

    #'''
    print('s', 'supData1', 'weight', 'supData1_w', 'supData2')
    for ls in list_freq:
        for s in ls:

            #1 按照apriori算法计算带权重的支持度
            # 当前项集的原始支持度
            supData1 = ss.map(lambda x: 1 if s.issubset(x) else 0).sum()
            # 当前项集中各元素权重的最小值
            weight = fp.df_weight[fp.df_weight[
                    'feature_group_name'].isin(s)]['weight'].min()
            # 当前项集的带权重支持度
            supData1_w = supData1 * weight

            #2 根据fpGrowth算法计算的带权重支持度
            # 当前项集利用fpGrowth计算带权重的的支持度
            xList, supData2 = fp.computeSupData(s.copy(), fpTree, dict_header)

            '''
            # 当前项集中各元素的支持度dict_sup和权重dict_weight
            #func_sup = lambda y: ss.map(lambda x: 1 if y.issubset(x) else 0).sum()
            #dict_sup = {x:func_sup({x}) for x in s}
            dict_weight = fp.df_weight[fp.df_weight['feature_group_name'].isin(s)
                    ].set_index('feature_group_name').to_dict()['weight']
            #list_sup = [dict_sup[x]*dict_weight[x] for x in dict_sup.keys()]
            #print(s, supData1, min(list_sup), supData2)

            # 当前项集的按fpGrowth计算的带权重支持度
            d = {x:dict_header[x][0] for x in s}
            # 降序(计数, 名称)排列
            list_ordered = [v[0] for v in sorted(d.items(), 
                    key=lambda x: (x[1], x[0]), reverse=True)]
            list_weighted = [dict_weight[x] for x in list_ordered]
            #print(d, list_ordered, list_weighted)
            # 取按这种排序方式得到的最末尾权重
            # weight*supData2/list_weighted[-1]即为转换成了
            # 按apriori方法计算的带权重支持度
            # supData1_w==weight*supData2/list_weighted[-1])
            '''

            #3 比较两种方法对同一项集计算出的支持度
            if supData1_w != supData2:
                print(s, supData1, weight, supData1_w, supData2, )

    #'''

def transInterval(x):
    """
        被predictModel()调用
        用于转换interval列, 将区间字符串转换为区间格式, 
        将数字字符转换成数字格式, 原本就是字符串的不用转换
    """

#    print([x])
    # 转换Interval类型
    if ']' in x:
        l = x.replace('(','').replace(']','').split(',')
        return pd.Interval(eval(l[0]), eval(l[1]))
    # 转换数字类型
    try:
        return eval(x)
    # 不处理字符串类型
    except:
        return x

def predictModel(ss_sample, df_std, ss_rules):
    """预测新样本的是否命中关注规则"""

    #print(ss_sample.tolist())
    #print(df_std.head())

    #1 去除做预测没用的指标
    for i in ['enterprise_id', 'alert_level']:
        del ss_sample[i]

    #2 将interval列(读入时均是字符串)转换成区间或数字格式
    # transInterval()后混合有pd.Interval, float和str
    df_std['Interval'] = df_std['interval'].map(transInterval)

    #3 依次判断待预测样本的每个指标属于什么分组
    # 初始化新样本每个指标的分组值列表
    l = []
    # 对新样本的每个指标
    #'''
    for i in ss_sample.index:
        #print(i, ss_sample[i])
        # 通过指标名筛选df_std, 先假设Interval列是区间
        # 尝试判断指标值是否在df_sel['Interval']中, 来筛选记录
        try:
            df_sel = df_std[df_std['feature']==i]
            df_sel = df_sel[df_sel['Interval'].map(lambda x: ss_sample[i] in x)]
        # 一旦报错, 则利用是否等于来筛选记录
        except:
            df_sel = df_std[(df_std['feature']==i)&(df_std['Interval']==ss_sample[i])]
        #print(df_sel)
        #print('\n')
        # 获取筛选记录的指标分组值
        l.append(df_sel['feature_group_name'].iloc[0])
    #print(len(l), '\n', l)
    # 构造新样本指标转换后的分组值
    ss_sample_trans = pd.Series(l)
    #'''

    #4 按照前件规则中的指标项组合, 判断待预测样本的命中情况
    # 只要命中了一条关注规则都需关注
    ss_hit = ss_rules.map(lambda x: [1 if i in l else 0 for i in x.split(';')])
    flag = any(ss_hit.map(all).tolist())
    # 构造df_hit是为了返回给testModule()作展示用, 下面返回df_std也是同样目的
    df_hit = pd.DataFrame({'关注规则': ss_rules, '命中情况': ss_hit})
    # 实际预测时, 只需要返回该样本是否命中的结果flag即可
    #flag = any(df_hit['命中情况'].map(all).tolist())
    #print(flag)

    # 返回是否需要关注, 用于构建重点关注名单
    return ss_sample_trans, df_std, df_hit, flag

def modelEvaluation(ss_real, ss_predict):
    """在测试结果上评估模型, 计算TP,TN,FP,FN,精确度,召回率"""

    #print(ss_real, ss_predict)
    df = pd.DataFrame({'real': ss_real, 'predict': ss_predict})
    TP = df[(df['real']==True)&(df['predict']==True)].shape[0]
    TN = df[(df['real']==False)&(df['predict']==False)].shape[0]
    FP = df[(df['real']==False)&(df['predict']==True)].shape[0]
    FN = df[(df['real']==True)&(df['predict']==False)].shape[0]
    Accuracy = (TP + TN) / (TP + TN + FP + FN)
    Precision = round(TP / (TP + FP), 4)
    Recall = round(TP / (TP + FN), 4)
    F1Score = round(2*Precision*Recall/(Precision+Recall), 4)

    return TP, TN, FP, FN, Accuracy, Precision, Recall, F1Score

def showProcess(df_raw, df_train, df_test, minSup, minConf, init_weight):
    """按照demo展示流程, 输出建模关键处理结果"""

    #1 获取样本数据
    #df_raw = pd.read_csv('datasets/features.csv', dtype={'enterprise_id':str})
    print('#1 获取样本数据:\n')
    print('#1 总的样本数, 特征数:', df_raw.shape)
    print('#1 前几个样本值举例展示:\n', df_raw.head()) 
    print('#1 其中的特征名有:\n', df_raw.columns)
    print('\n')

    #2 划分训练和测试样本集, 训练样本和测试样本数量比为9:1
    #df_test = df_raw.sample(frac=0.1, replace=False, axis=0, random_state=1) 
    #df_train = df_raw[~df_raw['enterprise_id'].isin(df_test['enterprise_id'])]
    print('#2 划分训练和训练样本集:\n')
    print('#2 训练样本集形状:', df_train.shape)
    print('#2 训练样本集的前几个举例展示:\n', df_train.head())
    print('#2 测试样本集形状:', df_test.shape)
    print('#2 测试样本集的前几个举例展示:\n', df_test.head())
    print('\n')

    # 设置最小支持度阈值, 如果传入的是小数, 则认为是占比, 将其转换为数量
    #minSup = 0.7
    #minConf = 0.8
    #init_weight = True

    #3 开始训练模型
    #3.0 初始化模型处理类, 传入训练数据
    fp = fpGrowth(df_train, minSup, minConf, init_weight)
    #3.1 模型训练第一步: 特征转换
    datasets_init = fp.loadAndTransDatasetsReal()
    fp.loadWeightReal()
    datasets = fp.initDatasets(datasets_init)

    list_features_trans = ['industry', 'registered_capital']
    print('#3.1 模型训练第一步: 特征转换\n')
    print('#3.1 以其中两个指标举例展示转换情况\n')
    print('#3.1 特征转换前:\n', df_train[list_features_trans].head())

    list_features_index = [df_train.columns.tolist().index(x) 
            for x in list_features_trans]
    print('#3.1 特征转换后:\n', pd.DataFrame(
                datasets_init[:5]).iloc[:, list_features_index])
    print('\n')

    #3.2 模型训练第二步: 人工确定各分组指标值的权重, 并读入
    if fp.init_weight:
        print('#3.2 各分组指标值权重(最后一列):\n', fp.df_weight)
        print('\n')

    #3.3 模型训练第三步: 查找频繁项集
    # 是否给项集添加权重
    dict_freq_sup = {}
    fpTree, dict_header = fp.createTree(datasets, {}, dict_freq_sup)
    print('\n#1 根据数据集构建的FP树:')
    fpTree.disp()
    print('\n#1 与FP树对应的头指针表:')
    print('单元素项集, 总计数值')
    for i,j in dict_header.items():
        print(i, j[0])

    #2 获取频繁项集
    fp.mineTree(fpTree, dict_header, set([]), dict_freq_sup)
    list_freq_all = dict_freq_sup.keys()
    list_freq = []
    for i in range(max([len(x) for x in list_freq_all])):
        list_freq.append([x for x in list_freq_all if len(x)==i+1])
    #print('包含所有项集的列表:', list_freq_all)
    print('\n#3.3 频繁项集的查找结果总览:')
    print('#3.3 频繁项集数:', len(list_freq_all))
    print('#3.3 不同元素数组合成的频繁项集的个数: ', [len(x) for x in list_freq])
    print('\n')

    #3.3 举例展示, 不同元素数组成的候选项集和频繁项集, 及其支持度
    #3.3 针对单元素
    df_l0 = pd.DataFrame({'频繁项集': list_freq[0]})
    df_l0['支持度'] = df_l0['频繁项集'].map(lambda x: dict_freq_sup[x])
            #fp.computeSupData(x.copy(), fpTree, dict_header)[1])
    print('#3.3 前5个单元素频繁项集及其支持度:\n', df_l0.head())
    print('\n')
    #3.3 针对二元素
    df_l1 = pd.DataFrame({'频繁项集': list_freq[1]})
    df_l1['支持度'] = df_l1['频繁项集'].map(lambda x: dict_freq_sup[x])
            #lambda x: fp.computeSupData(x.copy(), fpTree, dict_header)[1])
    print('#3.3 前5个二元素频繁项集及其支持度:\n', df_l1.head())
    print('\n')

    #3.4 模型训练第四步: 关联规则, 输出关联规则的置信度
    fp.dict_freq_sup = dict_freq_sup
    list_rules_all = fp.generateRules(list_freq, fpTree, dict_header)
    print('#3.4 从不同元素数的频繁项集中发现的关联规则数,')
    print('#3.4 分别为由2个以上元素构成的频繁项集发现的关联规则数:')
    rules_count = [len(l) for l in list_rules_all]
    print('\n#3 关联规则数:', sum(rules_count), rules_count)
    print('\n')

    print('#3.4 举例展示, 从两个元素的频繁项集中发现的关联规则(前5个):')
    for i in list_rules_all[0][:5]:
        print('置信度:', i[2])
        print('前件规则:\n', i[0])
        print('----> 后件规则:\n', i[1], '\n')
    print('\n')
    print('#3.4 举例展示, 从三个元素的频繁项集中发现的关联规则(前5个):')
    for i in list_rules_all[1][:5]:
        print('置信度:', i[2])
        print('前件规则:\n', i[0])
        print('----> 后件规则:\n', i[1], '\n')
    print('\n')

    #3.5 模型训练第五步: 筛选关注规则
    #3.5 挑选后件规则集合中存在关注级别较高的4或5级, 作为关注关联规则
    list_rules_sel = [x for y in list_rules_all for x in y 
            if 'alert_level__5' in x[1] or 'alert_level__4' in x[1]]
    print('#4 关注规则数:', len(list_rules_sel))
    print('#4 举例展示, 筛选关注等级较高的关联规则作为关注规则:')
    for i in list_rules_sel[:5]:
        print('置信度:', i[2])
        print('前件规则:\n', i[0])
        print('----> 后件规则:\n', i[1], '\n')
    print('\n')

    #4 存储模型结果
    filename = 'datasets/rules.xlsx'
    fp.saveModel(filename, list_freq, fpTree, dict_header, 
            list_rules_all, list_rules_sel)
    print('\n模型结果已保存')

    #5 利用模型结果预测新样本
    #5.1 导入分组标准, 从指标分组值(名)中获取指标名
    df_std = pd.read_excel('datasets/features_group_name.xlsx')
    df_std['feature'] = df_std['feature_group_name'].map(lambda x: x.split('__')[0])
    #5.3 导入关注规则, 前件规则去重
    df_rules = pd.read_excel('datasets/rules.xlsx', sheet_name='需关注的关联规则')
    ss_rules = df_rules['前件规则'].drop_duplicates()
    print('#5 展示几个新样本的预测情况:\n')
    for i in range(df_test.shape[0])[:2]:
        ss = df_test.iloc[i, :]
        print('#5 第%s个待预测样本信息:\n'% (i+1), ss)
        ss_sample_trans, df_std, df_hit, flag = predictModel(ss, df_std, ss_rules)
        print('\n#5 转换后的样本信息(用于匹配关注前件规则, '+\
                    '去除了enterprise_id和alert_level):\n', ss_sample_trans)
        print('\n#5 去重后的关注规则(即前件规则)的命中情况:\n', df_hit)
        print('\n#5 第%s个待预测样本的最终的命中情况(即是否关注):'% (i+1), flag)
        print('\n')

    '''
    #6 模型评估, 准确度和稳定性
    #print(df_test[:5]['alert_level'])
    print('#6 模型评估:')
    print('#6 待预测样本数:', df_test.shape)
    ss_real = df_test[:100]['alert_level'].map(
            lambda x: True if x in [4, 5] else False)
    ss_predict = df_test[:100].apply(
            lambda ss: predictModel(ss, df_std, ss_rules)[-1], axis=1)
    print('#6 测试样本的实际标签(展示前100个):\n', ss_real)
    print('#6 测试样本的预测标签(展示前100个):\n', ss_predict)
    TP, TN, FP, FN, Accuracy, Precision, Recall, F1Score = modelEvaluation(ss_real, ss_predict)
    print('\n')
    print('#6 TP, TN, FP, FN:', [TP, TN, FP, FN])
    print('#6 accuracy (准确率): ', Accuracy)
    print('#6 precision (精确率, 查准率): ', Precision)
    print('#6 recall (召回率, 查全率): ', Recall)
    '''


if __name__ == "__main__":

    #0 读入原始数据集, 切分训练集和测试集
    df_raw = pd.read_csv('datasets/features.csv', dtype={'enterprise_id': str})
    #print(df_raw.shape, '\n', df_raw.head()) 
    df_test = df_raw.sample(frac=0.1, replace=False, axis=0, random_state=1) 
    df_train = df_raw[~df_raw['enterprise_id'].isin(df_test['enterprise_id'])]

    # 设置最小支持度阈值minSup, 关联规则最小置信度minConf, 
    # 是否增加项集权重init_weight, 
    # 是否采用小样本手动构造测试数据集sampleOrReal, False代表采用随机生成的大数据集
    minSup = 0.7
    minConf = 0.9
    init_weight = True
    sampleOrReal = False

    # 训练模型, 保存模型结果, 模型测试
    trainModel(df_train, minSup, minConf, init_weight, sampleOrReal)

    #testComputingSupData(df_train, minSup, minConf, init_weight)

    # demo展示用函数
    #showProcess(df_raw, df_train, df_test, minSup, minConf, init_weight)


